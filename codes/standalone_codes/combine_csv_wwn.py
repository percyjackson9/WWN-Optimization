'''
Irrelevent code for the main algorithm.

'''


import pandas as pd
import glob
from openpyxl import load_workbook
import os
import pickle


with open("regina_flowclus_instances.pykl","rb") as f:
    task_regina=pickle.load(f)

# with open("LA_instances.pykl","rb") as f:
#     task_LA=pickle.load(f)


bd_regina=pd.DataFrame()
for i in task_regina.keys():
    if len(task_regina[i]) <= 3:
        name = "regina_" + "_".join([str(kk) for kk in task_regina[i]])
    else:
        name=f"regina_{task_regina[i][0]}_{task_regina[i][1]}_{task_regina[i][2]}_{task_regina[i][4]}"

    csv_name=f"./ResultsReginaClus2/{name}.csv"
    try:
        dt=pd.read_csv(csv_name)
        print(f"File for {i}:{name} found and read correctly")
        if len(task_regina[i]) > 3:
            dt["NOS"]=[task_regina[i][2]]
            dt["Strategy"]=[task_regina[i][0]]
            dt["NOC"]=[task_regina[i][3]]
            dt["Target"]=[task_regina[i][4]]

        else:
            dt["NOS"] = [task_regina[i][2]]
            dt["Strategy"] = [task_regina[i][0]]
    except:
        print(f"File not found, instance not solved, {i}:{name}")
        if len(task_regina[i])>3:
            data = {"NOS": [task_regina[i][2]],
                    "Strategy": [task_regina[i][0]],
                    "NOC":[task_regina[i][3]],
                    "Target":[task_regina[i][4]]
                    }

        else:
            data = {"NOS": [task_regina[i][2]],
                "Strategy": [task_regina[i][0]]
                }
        dt=pd.DataFrame(data)

    bd_regina = pd.concat([bd_regina, dt], ignore_index=True)

bd_regina.to_csv("./ResultsReginaClus2/regina_clus_results.csv")
#
# bd_LA = pd.DataFrame()
# for i in task_LA.keys():
#     if len(task_LA[i])<=3:
#         name = "LA_" + "_".join([str(kk) for kk in task_LA[i]])
#     else:
#         name=f"LA_{task_LA[i][0]}_{task_LA[i][1]}_{task_LA[i][2]}_{task_LA[i][4]}"
#     csv_name = f"./ResultsLatest/{name}.csv"
#     try:
#         dt = pd.read_csv(csv_name)
#         print(f"File for {i}:{name} found and read correctly")
#         if len(task_LA[i]) > 3:
#             dt["NOS"]=[task_LA[i][2]]
#             dt["Strategy"]=[task_LA[i][0]]
#             dt["NOC"]=[task_LA[i][3]]
#             dt["Target"]=[task_LA[i][4]]
#
#         else:
#             dt["NOS"] = [task_LA[i][2]]
#             dt["Strategy"] = [task_LA[i][0]]
#     except:
#         print(f"File not found, instance not solved, {i}:{name}")
#         if len(task_LA[i]) > 3:
#             data = {"NOS": [task_LA[i][2]],
#                     "Strategy": [task_LA[i][0]],
#                     "NOC": [task_LA[i][3]],
#                     "Target": [task_LA[i][4]]
#                     }
#
#         else:
#             data = {"NOS": [task_LA[i][2]],
#                     "Strategy": [task_LA[i][0]]
#                     }
#         dt = pd.DataFrame(data)
#
#     bd_LA = pd.concat([bd_LA, dt], ignore_index=True)
#
# bd_LA.to_csv("./ResultsLatest/LA_results.csv")

#
# def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
#                        truncate_sheet=False,
#                        **to_excel_kwargs):
#     """
#     Append a DataFrame [df] to existing Excel file [filename]
#     into [sheet_name] Sheet.
#     If [filename] doesn't exist, then this function will create it.
#
#     @param filename: File path or existing ExcelWriter
#                      (Example: '/path/to/file.xlsx')
#     @param df: DataFrame to save to workbook
#     @param sheet_name: Name of sheet which will contain DataFrame.
#                        (default: 'Sheet1')
#     @param startrow: upper left cell row to dump data frame.
#                      Per default (startrow=None) calculate the last row
#                      in the existing DF and write to the next row...
#     @param truncate_sheet: truncate (remove and recreate) [sheet_name]
#                            before writing DataFrame to Excel file
#     @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
#                             [can be a dictionary]
#     @return: None
#
#     Usage examples:
#
#     >>> append_df_to_excel('d:/temp/test.xlsx', df)
#
#     >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)
#
#     >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
#                            index=False)
#
#     >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
#                            index=False, startrow=25)
#
#     (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
#     """
#     # Excel file doesn't exist - saving and exiting
#     if not os.path.isfile(filename):
#         df.to_excel(
#             filename,
#             sheet_name=sheet_name,
#             startrow=startrow if startrow is not None else 0,
#             **to_excel_kwargs)
#         return
#
#     # ignore [engine] parameter if it was passed
#     if 'engine' in to_excel_kwargs:
#         to_excel_kwargs.pop('engine')
#
#     writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')
#
#     # try to open an existing workbook
#     writer.book = load_workbook(filename)
#
#     # get the last row in the existing Excel sheet
#     # if it was not specified explicitly
#     if startrow is None and sheet_name in writer.book.sheetnames:
#         startrow = writer.book[sheet_name].max_row
#
#     # truncate sheet
#     if truncate_sheet and sheet_name in writer.book.sheetnames:
#         # index of [sheet_name] sheet
#         idx = writer.book.sheetnames.index(sheet_name)
#         # remove [sheet_name]
#         writer.book.remove(writer.book.worksheets[idx])
#         # create an empty sheet [sheet_name] using old index
#         writer.book.create_sheet(sheet_name, idx)
#
#     # copy existing sheets
#     writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
#
#     if startrow is None:
#         startrow = 0
#
#     # write out the new sheet
#     df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
#
#     # save the workbook
#     writer.save()
#
# for OID in obst:
#     for NAMES in names:
#         print(f"Writing for {OID} - {NAMES}")
#         #nodes = [300,500,700,1000]#, 500, 700, 1000, 2000]
#
#         nodes=[30,50,100,200,300,500,1000]
#         replications = [1, 2, 3, 4, 5]
#         p = [5, 10, 15]  ## no of relays
#
#         #cgs=[8,10]
#         name=f"./PaperFinalNov/{OID}/R_{NAMES}/"
#         sname=f"{NAMES}"
#         filename_excel = f"./PaperFinalNov/RNPPO_{OID}_results.xlsx"
#         if "imp" in sname:
#             full_data=pd.DataFrame(columns=["NS","IMP","NK","DS","ID","OBJ","TIME","TICKS","GAP",
#                                           "ERR", "STATUS","DATE", "TOTH",
#                                           "SUCH", "FREQ","WOBJ","WT",
#                                           "LOCS","SEL","CONN"])
#         else:
#             full_data = pd.DataFrame(columns=["NS","NK","DS","ID","OBJ","TIME","TICKS","GAP",
#                                           "ERR", "STATUS","DATE", "TOTH",
#                                           "SUCH", "FREQ","WOBJ","WT",
#                                           "LOCS","SEL","CONN"])
#         for n in nodes:
#             if n<=200:
#                 cgs = [10, 15, 20]  ## sensor radius
#             else:
#                 cgs = [8,10]  ## sensor radius
#             for prob in p:
#                 for cg in cgs:
#                     for reps in replications:
#                         fname = name + f"{OID}-{n}-{prob}-{cg}-{reps}-{NAMES}*.csv"
#                         try:
#                             files=glob.glob(fname)
#                         except:
#                             print("File/Folder not found")
#                             continue
#
#
#                         if len(files)>0:
#                             print(f"Instance result for {OID}-{n}-{prob}-{cg}-{reps}-{NAMES} present")
#                             filename=files[0]
#                             data=pd.read_csv(filename)
#
#                             accepted_columns = ["NS","NK","DS","ID","Obj","Gap","Time","Ticks","Error","Status","Date",
#                                                 "Total Heur","Succ Heur","Freq","Locs","Selected","Conn","WObj","WTime"]
#                             to_delete = []
#                             for i in data.columns:
#                                 if i not in accepted_columns:
#                                     to_delete.append(i)
#
#                             data = data.drop(to_delete, axis=1)
#
#
#                             if "imp" not in NAMES:
#                                 new_data={"NS":[n],"NK":[prob],"DS":[cg],"ID":[reps],
#                                           "OBJ":[data.loc[0,"Obj"]],"TIME":[data.loc[0,"Time"]],"TICKS":[data.loc[0,"Ticks"]],"GAP":[data.loc[0,"Gap"]],
#                                           "ERR": [data.loc[0, "Error"]], "STATUS": [data.loc[0, "Status"]],"DATE": [data.loc[0, "Date"]], "TOTH": [data.loc[0, "Total Heur"]],
#                                           "SUCH": [data.loc[0, "Succ Heur"]], "FREQ": [data.loc[0, "Freq"]],
#                                           "LOCS": [data.loc[0, "Locs"]],"SEL": [data.loc[0, "Selected"]],"CONN":[data.loc[0, "Conn"]]
#                                       }
#                             else:
#                                 new_data = {"NS": [n], "IMP":[data.loc[0, "NS"]], "NK": [prob], "DS": [cg], "ID": [reps],
#                                             "OBJ": [data.loc[0, "Obj"]], "TIME": [data.loc[0, "Time"]],
#                                             "TICKS": [data.loc[0, "Ticks"]], "GAP": [data.loc[0, "Gap"]],
#                                             "ERR": [data.loc[0, "Error"]], "STATUS": [data.loc[0, "Status"]],
#                                             "DATE": [data.loc[0, "Date"]], "TOTH": [data.loc[0, "Total Heur"]],
#                                             "SUCH": [data.loc[0, "Succ Heur"]], "FREQ": [data.loc[0, "Freq"]],
#                                             "LOCS": [data.loc[0, "Locs"]], "SEL": [data.loc[0, "Selected"]],"CONN":[data.loc[0, "Conn"]]}
#
#                             if "warm" in NAMES:
#                                 new_data["WOBJ"]=[data.loc[0,"WObj"]]
#                                 new_data["WT"]=[data.loc[0,"WTime"]]
#                             else:
#                                 new_data["WOBJ"] = [0]
#                                 new_data["WT"] = [0]
#
#
#                             new_data=pd.DataFrame(new_data)
#                             full_data = pd.concat([full_data, new_data], ignore_index=True)
#
#                         else:
#                             print(f"Instance result for {OID}-{n}-{prob}-{cg}-{reps}-{NAMES} absent")
#
#                             if "imp" not in NAMES:
#                                 new_data = {"NS": [n], "NK": [prob], "DS": [cg], "ID": [reps],
#                                             "OBJ": [0], "TIME": [0],
#                                             "TICKS": [0], "GAP": [0],
#                                             "ERR": [0], "STATUS": ["INF"],
#                                             "DATE": [0], "TOTH": [0],
#                                             "SUCH": [0], "FREQ": [0],
#                                             "LOCS": [0], "SEL": [0],"CONN":[0],"WOBJ":[0],"WT":[0]
#                                             }
#                             else:
#                                 new_data = {"NS": [n],"IMP":[0], "NK": [prob], "DS": [cg], "ID": [reps],
#                                             "OBJ": [0], "TIME": [0],
#                                             "TICKS": [0], "GAP": [0],
#                                             "ERR": [0], "STATUS": ["INF"],
#                                             "DATE": [0], "TOTH": [0],
#                                             "SUCH": [0], "FREQ": [0],
#                                             "LOCS": [0], "SEL": [0],"CONN":[0],"WOBJ":[0],"WT":[0]
#                                             }
#
#                             new_data = pd.DataFrame(new_data)
#                             #df_dictionary = pd.DataFrame([new_data])
#                             full_data = pd.concat([full_data, new_data], ignore_index=True)
#
#         full_data.to_csv(f"./PaperFinalNov/RNPPO_{OID}_{sname}_results.csv")
#         GFG = pd.ExcelWriter(filename_excel,engine="openpyxl",if_sheet_exists="replace",mode="a")
#         GFG.book = load_workbook(filename_excel)
#         # copy existing sheets
#         GFG.sheets = dict((ws.title, ws) for ws in GFG.book.worksheets)
#         full_data.to_excel(GFG,sheet_name=NAMES)
#         GFG.save()